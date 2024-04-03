#!/usr/bin/env python3
from __future__ import annotations

import json
import logging

from os.path import exists as file_exists
from typing import Any
from openpyxl import Workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule


logger = logging.getLogger("autovmaf-excel")
JOBNAMES_FILE = 'jobnames.txt'

if not file_exists(JOBNAMES_FILE):
    exit(f'Could not find "{JOBNAMES_FILE}"')

with open(JOBNAMES_FILE, encoding="utf-8") as f:
    # remove comments, then filter empty lines
    jobnames = [line.split("#")[0].strip() for line in f]
    jobnames = [*filter(None, jobnames)]

wb = Workbook()

redLine = 'ff0000'
redFill = PatternFill(start_color='ff9494',
                      end_color='ff9494',
                      fill_type='solid')

orangeFill = PatternFill(start_color='ffcf94',
                         end_color='ffcf94',
                         fill_type='solid')

greenFill = PatternFill(start_color='d3eeb4',
                        end_color='d3eeb4',
                        fill_type='solid')

border_style = Side(border_style="medium", color="336a15")
border = Border(top=border_style, bottom=border_style,
                left=border_style, right=border_style)


for jobname in jobnames:
    RESULT_FILE = f"./results/{jobname}.json"
    if not file_exists(RESULT_FILE):
        logger.warning(f'⚠️  Skipping "{RESULT_FILE}": Not found.')
        continue
    try:
        with open(RESULT_FILE, encoding="utf-8") as f:
            json_data = json.load(f)
            result: dict[str, Any] = json_data["result"]
            models = result.get(jobname)
            if not models:
                result_names = [*result.keys()]
                if len(result_names) == 1:
                    logger.warning(f'⚠️  "{RESULT_FILE}" does not contain the result for "{jobname}", using "{result_names[0]}" instead. This can happen if renaming the file.')
                    models = result[result_names[0]]
                else:
                    logger.warning(f'⚠️  "{RESULT_FILE}" results "{result_names}" does not contain "{jobname}"')
                    continue
    except Exception as err:
        logger.warning(f'⚠️  Skipping "{RESULT_FILE}": Could not parse.', err)
        continue


    jobname_trunc = jobname[:31]
    heights = []
    bitrates = []
    resolutions = []

    for model in models:
        results = models[model]

        if len(results) > 0:
            r = []
            for entry in results:
                e = {}
                res, bitrate = entry.split("_")[:2]
                e["res"] = res
                width, height = res.split("x")
                e["width"] = int(width)
                e["height"] = int(height)
                e["bitrate"] = int(bitrate)
                e["model"] = model
                e["score"] = float(results[entry])
                e["jobname"] = jobname
                heights.append(int(height))
                bitrates.append(int(bitrate))
                resolutions.append(res)
                r.append(e)

            scores = sorted(r, key=lambda i: (i["height"], i["bitrate"]))
            heights = sorted(list(set(heights)))
            bitrates = sorted(list(set(bitrates)))
            resolutions = sorted(
                list(set(resolutions)), key=lambda x: int(x.split("x")[1])
            )

    if wb.sheetnames[0] == "Sheet":
        wb.remove(wb.active)

    ws = wb.create_sheet(jobname_trunc)

    table = [["Kbit/s", "Ref"] + resolutions]

    for b in bitrates:
        # 93 is the reference point
        row = [b/1000, 93]
        for h in heights:
            matches = [
                item
                for item in scores
                if (item["bitrate"] == b and item["height"] == h)
            ]

            if len(matches) > 0 and matches[0]["bitrate"] == b:
                row.append(matches[0]["score"])
            else:
                row.append("")
        table.append(row)

    for row in table:
        ws.append(row)

    COLUMN_LIST = [get_column_letter(cell.column) for cell in ws[1]]
    RANGE_STRING = f'{jobname_trunc}!A1:{COLUMN_LIST[-1]}{str(len(table))}'
    OFFSET = 10
    ws.move_range(RANGE_STRING, rows=0, cols=OFFSET)

    MIN_COL = 2 + OFFSET
    MIN_ROW = 1
    MAX_COL = len(resolutions) + 2 + OFFSET
    MAX_ROW = len(table)

    values = Reference(
        ws, min_col=MIN_COL, min_row=MIN_ROW, max_col=MAX_COL, max_row=MAX_ROW
    )

    OFFSET_COL = chr(ord('A') + OFFSET)
    RANGE_STRING = f'{jobname_trunc}!{OFFSET_COL}2:{OFFSET_COL}{str(len(table))}'
    X_VALUES = Reference(ws, range_string=RANGE_STRING)

    chart = LineChart()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(X_VALUES)

    # cosmetics
    chart.title = jobname
    chart.height = 17
    chart.width = 25
    chart.style = 2
    chart.legend.position = "b"
    chart.x_axis.title = "Bitrate in Kbit/s"
    chart.y_axis.title = "VMAF Score"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100

    # Red line for constant reference value
    chart_data_series = chart.series[0]
    chart_data_series.graphicalProperties.line.width = pixels_to_EMU(4)
    chart_data_series.graphicalProperties.line.solidFill = redLine
    chart_data_series.marker.graphicalProperties.solidFill = redLine
    chart_data_series.marker.symbol = "diamond"
    chart_data_series.marker.size = 2

    ws.add_chart(chart, "A1")
    ACTIVE_COLUMN_LIST = [get_column_letter(cell.column) for cell in ws[1]]

    for letter in ACTIVE_COLUMN_LIST:
        ws.column_dimensions[letter].width = 12

    NEW_START_COL = chr(ord('A') + OFFSET + 1)
    DATA_RANGE = f'{NEW_START_COL}2:{ACTIVE_COLUMN_LIST[-1]}{str(len(table))}'

    ws.conditional_formatting.add(DATA_RANGE,
                                  CellIsRule(
                                      operator='between',
                                      formula=[93.0, 94.0],
                                      stopIfTrue=True,
                                      fill=orangeFill))
    ws.conditional_formatting.add(DATA_RANGE,
                                  CellIsRule(
                                      operator='between',
                                      formula=[0.1, 20],
                                      stopIfTrue=True,
                                      fill=orangeFill))
    ws.conditional_formatting.add(DATA_RANGE,
                                  CellIsRule(
                                      operator='greaterThanOrEqual',
                                      formula=[94.0],
                                      stopIfTrue=True,
                                      fill=redFill))

    LADDER_FILE = f"./results/{jobname}_ladder.json"

    if file_exists(LADDER_FILE):
        with open(LADDER_FILE, encoding="utf-8") as f:
            try:
                the_ladder = json.load(f)
            except json.decoder.JSONDecodeError:
                logger.warning(f"⚠️  Can't parse {LADDER_FILE}")

        autoladder = [["Auto-Ladder", "", ""]]

        try:
            if the_ladder:
                for ladder in the_ladder:
                    w = ladder["resolution"]["width"]
                    h = ladder["resolution"]["height"]

                    row = [f'{w}x{h}', ladder["bitrate"]/1000, ladder["vmaf"]]
                    autoladder.append(row)
            else:
                logger.warning(f'⚠️  Can not generate auto-ladder for resolution {w}x{h} bw:{ladder["bitrate"]/1000}')

            scores = [s[2] for s in autoladder]

            for row in ws.rows:
                for cell in row:
                    if cell.value is not (None or "") and cell.value in scores:
                        cell.fill = greenFill
                        cell.border = border

            for row in autoladder:
                ws.append(row)
        except NameError:
            pass
    else:
        logger.warning(f'⚠️  Can not find {LADDER_FILE}, will not generate auto-ladder')

wb.save("vmaf.xlsx")
print("File generated!")
